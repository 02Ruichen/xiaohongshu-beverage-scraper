"""
给增强版 Excel 追加数据透视 Sheet
"""
import pandas as pd
import io, sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

SRC = Path("D:/claude/AI race forest/output/饮料赛道_小红书_增强版_20260714.xlsx")

# 颜色 & 样式
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
TITLE_FONT = Font(name="微软雅黑", size=13, bold=True, color="1F3864")
ACCENT_FILL = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")   # 浅紫
LIGHT_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_ORANGE = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

GROUP_ORDER = ["🍹 口味上新", "💚 健康诉求", "📸 视觉圈层", "⚠️ 踩雷避坑"]
GROUP_COLORS = {
    "🍹 口味上新": LIGHT_ORANGE,
    "💚 健康诉求": LIGHT_GREEN,
    "📸 视觉圈层": LIGHT_BLUE,
    "⚠️ 踩雷避坑": ACCENT_FILL,
}

print("📖 读取增强版数据...")
wb = load_workbook(SRC)

# 读出现有数据
notes = pd.read_excel(SRC, sheet_name="笔记列表")
comments = pd.read_excel(SRC, sheet_name="评论明细")

# ════════════════════════════════════════════════════════
# 辅助函数
# ════════════════════════════════════════════════════════
def style_header(ws, row, ncols):
    for ci in range(1, ncols + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data_row(ws, row, ncols, bold=False, fill=None):
    for ci in range(1, ncols + 1):
        cell = ws.cell(row=row, column=ci)
        cell.font = BOLD_FONT if bold else DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if fill:
            cell.fill = fill

def auto_width(ws, widths: list):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def write_matrix(ws, start_row: int, headers: list, row_labels: list, data_matrix: list[list],
                 label_width: int = 20, col_widths: list = None, row_fills: list = None,
                 total_row: bool = True):
    """往 worksheet 写入一个矩阵表格，返回结束行号"""
    ncols = len(headers) + 1
    r = start_row

    # 表头
    ws.cell(row=r, column=1, value="").fill = HEADER_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for ci, h in enumerate(headers, 2):
        ws.cell(row=r, column=ci, value=h)
    style_header(ws, r, ncols)
    r += 1

    # 数据行
    for ri, (label, values) in enumerate(zip(row_labels, data_matrix)):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = THIN_BORDER
        fill = row_fills[ri] if row_fills else None
        for ci, val in enumerate(values, 2):
            ws.cell(row=r, column=ci, value=val)
        style_data_row(ws, r, ncols, fill=fill)
        r += 1

    # 合计行
    if total_row and data_matrix:
        ws.cell(row=r, column=1, value="📊 合计")
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = THIN_BORDER
        for ci in range(2, ncols + 1):
            col_total = sum(row[ci-2] for row in data_matrix if isinstance(row[ci-2], (int, float)))
            ws.cell(row=r, column=ci, value=col_total)
            ws.cell(row=r, column=ci).fill = TOTAL_FILL
        style_data_row(ws, r, ncols, bold=True, fill=TOTAL_FILL)
        r += 1
    elif total_row:
        r += 1

    # 列宽
    ws.column_dimensions[get_column_letter(1)].width = label_width
    if col_widths:
        for ci, w in enumerate(col_widths, 2):
            ws.column_dimensions[get_column_letter(ci)].width = w
    return r

# ════════════════════════════════════════════════════════
# Sheet 5: 透视-组别×评论类型
# ════════════════════════════════════════════════════════
print("📊 生成 透视-组别×评论类型...")
ws1 = wb.create_sheet("透视-组别×评论类型")

# 构建交叉表
rows_data = []
for g in GROUP_ORDER:
    gn = notes[notes["组别"] == g]
    gc = comments[comments["组别"] == g]
    main_c = len(gc[gc["评论类型"] == "主评论"])
    sub_c = len(gc[gc["评论类型"] == "子级回复"])
    rows_data.append({
        "组别": g,
        "笔记数": len(gn),
        "评论总计": len(gc),
        "主评论": main_c,
        "子回复": sub_c,
        "平均点赞": round(gn["点赞数"].mean(), 0),
        "最高点赞": gn["点赞数"].max(),
        "最低点赞": gn["点赞数"].min(),
        "平均每笔记评论": round(len(gc) / len(gn), 1) if len(gn) > 0 else 0,
    })

headers = ["笔记数", "评论总计", "主评论", "子回复", "平均点赞", "最高点赞", "最低点赞", "平均每笔记评论"]
labels = [r["组别"] for r in rows_data]
matrix = [[r[h] for h in headers] for r in rows_data]
fills = [GROUP_COLORS.get(g, None) for g in labels]

end_r = write_matrix(ws1, 1, headers, labels, matrix,
                     label_width=18,
                     col_widths=[10, 10, 10, 10, 12, 12, 12, 14],
                     row_fills=fills)
ws1.freeze_panes = "A2"

# ════════════════════════════════════════════════════════
# Sheet 6: 透视-关键词明细
# ════════════════════════════════════════════════════════
print("📊 生成 透视-关键词明细...")
ws2 = wb.create_sheet("透视-关键词明细")

# 按关键词汇总
all_keywords = []
for g in GROUP_ORDER:
    g_kws = [k for k, v in {
        "奶茶新品":"🍹 口味上新","果茶推荐":"🍹 口味上新","咖啡新品":"🍹 口味上新",
        "低卡奶茶":"💚 健康诉求","无糖饮料推荐":"💚 健康诉求","减脂期喝什么":"💚 健康诉求",
        "奶茶探店":"📸 视觉圈层","夏日饮品合集":"📸 视觉圈层","网红奶茶打卡":"📸 视觉圈层",
        "奶茶避雷":"⚠️ 踩雷避坑","饮料踩雷":"⚠️ 踩雷避坑","最难喝饮料":"⚠️ 踩雷避坑",
    }.items() if v == g]
    for kw in g_kws:
        gn = notes[notes["来源关键词"] == kw]
        # 通过笔记标题找到对应评论
        note_titles = gn["标题"].tolist()
        gc = comments[comments["来源笔记"].isin(note_titles)]
        main_c = len(gc[gc["评论类型"] == "主评论"])
        sub_c = len(gc[gc["评论类型"] == "子级回复"])
        all_keywords.append({
            "组别": g,
            "关键词": kw,
            "笔记数": len(gn),
            "评论总计": len(gc),
            "主评论": main_c,
            "子回复": sub_c,
            "平均点赞": round(gn["点赞数"].mean(), 0),
            "最高点赞": gn["点赞数"].max(),
        })

kw_headers = ["组别", "笔记数", "评论总计", "主评论", "子回复", "平均点赞", "最高点赞"]
kw_labels = [f"{r['组别']}  |  {r['关键词']}" for r in all_keywords]
kw_matrix = [[r[h] for h in kw_headers] for r in all_keywords]
kw_fills = [GROUP_COLORS.get(r["组别"], None) for r in all_keywords]

end_r2 = write_matrix(ws2, 1, kw_headers, kw_labels, kw_matrix,
                      label_width=30,
                      col_widths=[8, 10, 10, 10, 10, 12, 12],
                      row_fills=kw_fills)
ws2.freeze_panes = "A2"

# ════════════════════════════════════════════════════════
# Sheet 7: 透视-TOP笔记排行
# ════════════════════════════════════════════════════════
print("📊 生成 透视-TOP笔记排行...")
ws3 = wb.create_sheet("透视-TOP笔记排行")

top_notes = notes.nlargest(15, "点赞数").copy()
top_notes["排名"] = range(1, len(top_notes) + 1)
top_notes["评论数"] = top_notes["标题"].apply(
    lambda t: len(comments[comments["来源笔记"] == t])
)

tn_headers = ["排名", "标题", "作者", "组别", "点赞数", "评论数", "来源关键词"]
tn_data = []
for _, r in top_notes.iterrows():
    tn_data.append([r[h] for h in tn_headers])

tn_fills = [GROUP_COLORS.get(r["组别"], None) for _, r in top_notes.iterrows()]

# 写表头
r = 1
for ci, h in enumerate(tn_headers, 1):
    ws3.cell(row=r, column=ci, value=h)
style_header(ws3, r, len(tn_headers))
r += 1

for ri, (row_data, fill) in enumerate(zip(tn_data, tn_fills)):
    for ci, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=ci, value=val)
    style_data_row(ws3, r, len(tn_headers), fill=fill)
    # 标题列左对齐
    ws3.cell(row=r, column=2).alignment = LEFT_WRAP
    r += 1

auto_width(ws3, [6, 50, 18, 18, 10, 8, 18])
ws3.freeze_panes = "A2"

# ════════════════════════════════════════════════════════
# Sheet 8: 透视-评论焦点 (主评论TOP + 各组热评)
# ════════════════════════════════════════════════════════
print("📊 生成 透视-评论焦点...")
ws4 = wb.create_sheet("透视-评论焦点")

# TOP 主评论（按点赞排序）
top_comments = comments[comments["评论类型"] == "主评论"].nlargest(20, "点赞数").copy()

r = 1
# 标题
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws4.cell(row=r, column=1, value="🔥 热度 TOP 20 主评论").font = TITLE_FONT
ws4.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
r += 2

tc_headers = ["排名", "组别", "评论文本", "点赞数", "来源笔记"]
for ci, h in enumerate(tc_headers, 1):
    ws4.cell(row=r, column=ci, value=h)
style_header(ws4, r, len(tc_headers))
r += 1

for i, (_, tc) in enumerate(top_comments.iterrows(), 1):
    vals = [i, tc["组别"], tc["评论文本"], tc["点赞数"], tc["来源笔记"]]
    for ci, val in enumerate(vals, 1):
        ws4.cell(row=r, column=ci, value=val)
    style_data_row(ws4, r, len(tc_headers), fill=GROUP_COLORS.get(tc["组别"], None))
    ws4.cell(row=r, column=3).alignment = LEFT_WRAP  # 评论文字左对齐
    r += 1

# 各组热评 TOP3
r += 2
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws4.cell(row=r, column=1, value="📋 各组热评 TOP 3").font = TITLE_FONT
r += 2

for gi, group_name in enumerate(GROUP_ORDER):
    gc = comments[(comments["组别"] == group_name) & (comments["评论类型"] == "主评论")].nlargest(3, "点赞数")
    if len(gc) == 0:
        continue

    # 组标题
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws4.cell(row=r, column=1, value=f"▍{group_name}").font = Font(name="微软雅黑", size=11, bold=True, color="1F3864")
    ws4.cell(row=r, column=1).fill = GROUP_COLORS.get(group_name, LIGHT_BLUE)
    for ci in range(1, 6):
        ws4.cell(row=r, column=ci).fill = GROUP_COLORS.get(group_name, LIGHT_BLUE)
    r += 1

    for _, row in gc.iterrows():
        vals = ["", "", row["评论文本"], row["点赞数"], row["来源笔记"]]
        for ci, val in enumerate(vals, 1):
            ws4.cell(row=r, column=ci, value=val)
        style_data_row(ws4, r, len(tc_headers))
        ws4.cell(row=r, column=3).alignment = LEFT_WRAP
        r += 1
    r += 1  # 组间空行

auto_width(ws4, [6, 18, 55, 8, 45])

# ════════════════════════════════════════════════════════
# 保存
# ════════════════════════════════════════════════════════
print("💾 保存中...")
wb.save(SRC)
print("🎉 数据透视 Sheet 全部添加完毕！")
print(f"   文件: {SRC}")
