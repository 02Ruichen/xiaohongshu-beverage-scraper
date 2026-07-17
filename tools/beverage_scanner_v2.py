"""
泛饮品赛道 —— 小红书数据抓取 v2（OpenCLI 轻量版）
=========================================================
核心改进（相比 Playwright 旧版）：
  ✅ 无需启动浏览器 — OpenCLI 复用 Chrome 登录态
  ✅ 无需反检测注入
  ✅ 评论自带楼中楼（--with-replies）
  ✅ 被封自动 fallback 到次热笔记
  ✅ 浅紫色标注主评论行
  ✅ 代码量减少 70%+

运行方式：
  python beverage_scanner_v2.py

前置条件：
  - opencli doctor 全绿（Chrome 扩展已连接）
  - Chrome 已登录 xiaohongshu.com（建议用小号）
  - 换号后第一次跑效果最好！

作者备注：诡秘专供 v2，边搜边抓防封吼~
"""

import json
import os
import subprocess
import sys
import time
import random
import re
import io
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

# 修复 Windows GBK 终端 emoji 乱码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# 项目根目录（tools/ 的上级）
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd


# ============================================================
# 🎛️  全局配置
# ============================================================

OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# ════════════════════════════════════════════
# 🎯 关键词组 → 从 config.py 读取
# ════════════════════════════════════════════
# 复制 config.example.py → config.py 后填入你的关键词
try:
    from config import KEYWORD_GROUPS, KW_TO_GROUP, GROUP_ORDER
except ImportError:
    print("❌ 找不到 config.py！")
    print("   请先复制 config.example.py → config.py 并填入你的关键词")
    sys.exit(1)

# 每关键词收录笔记数（12 × 4 = 48 条）
MAX_NOTES_PER_KEYWORD = 4

# 每笔记抓取评论数
MAX_COMMENTS_PER_NOTE = 30

# 关键词间休息（搜索阶段，搜完就过）
KEYWORD_DELAY_MIN = 3
KEYWORD_DELAY_MAX = 6

# 评论笔记间休息（集中抓评论阶段，间隔拉长防封）
COMMENT_NOTE_DELAY_MIN = 8
COMMENT_NOTE_DELAY_MAX = 16

# 评论抓取：全局热度 TOP N 条笔记
TOP_COMMENT_NOTES = 10

# 被风控后 fallback 到热度次高笔记
FALLBACK_DEPTH = 3


# ════════════════════════════════════════════
# 🧹 灌水词黑名单
# ════════════════════════════════════════════
NOISE_BLACKLIST = [
    "蹲", "蹲蹲", "蹲一个", "蹲一波",
    "马克", "m", "M", "码", "码住",
    "打卡", "滴", "滴滴", "ddd", "DDD", "dd",
    "顶", "顶顶", "1", "2", "3", "666", "999",
    "求链接", "求lj", "求个链接", "链接", "链接发我",
    "买", "买买买", "想买", "在哪买", "哪里买",
    "怎么买", "求购", "多少钱", "求店铺", "哪家",
    "有链接吗", "链接有吗",
    "接好运", "接接接", "接", "好运",
    "来了", "来了来了", "看看", "看看看",
    "前排", "沙发", "第一",
    "学到了", "收藏了", "收藏",
    "哈哈哈", "哈哈", "笑死", "笑晕",
    "好看", "好喝", "想要",
    "打卡打卡", "已阅",
]
RE_PURE_PUNCT = re.compile(r'^[\s.,;:!?。，、；：！？…～~\-—\-_#@￥$%^&*()（）、""''「」『』【】《》〈〉…"\']+$')
RE_PURE_DIGIT = re.compile(r'^\d+$')


# ============================================================
# 🏗️  数据结构
# ============================================================

@dataclass
class NoteCard:
    note_id: str = ""
    title: str = ""
    author: str = ""
    like_count: int = 0
    url: str = ""
    published_at: str = ""
    keyword_source: str = ""


@dataclass
class CommentRow:
    parent_id: str = ""
    ctype: str = ""           # "主评论" | "子级回复"
    text: str = ""
    likes: int = 0
    post_time: str = ""
    reply_to: str = ""
    note_url: str = ""
    note_title: str = ""


# ============================================================
# 🧰 工具函数
# ============================================================

def parse_count(text) -> int:
    """小红书数值 → 整数：支持 1.2万 / 3,456"""
    if not text:
        return 0
    if isinstance(text, (int, float)):
        return int(text)
    text = str(text).strip()
    if "万" in text:
        try:
            return int(float(text.replace("万", "").replace(",", "")) * 10000)
        except ValueError:
            return 0
    text = text.replace(",", "")
    try:
        return int(text)
    except ValueError:
        return 0


def is_noise(text: str) -> bool:
    """判断评论是否为灌水噪音"""
    if not text or len(text.strip()) < 5:
        return True
    t = text.strip()
    if RE_PURE_PUNCT.match(t) or RE_PURE_DIGIT.match(t):
        return True
    t_lower = t.lower()
    for nw in NOISE_BLACKLIST:
        if t_lower == nw.lower():
            return True
    if len(t) <= 2:
        for nw in NOISE_BLACKLIST:
            if nw in t:
                return True
    if re.match(r'^[求给发有要].{0,2}(链接|lj|🔗|地址|店铺|店).{0,2}$', t):
        return True
    return False


def run_opencli(args: list, timeout: int = 45) -> dict | list:
    """
    调用 OpenCLI，返回解析后的 JSON。
    自动处理 Windows .cmd 路径和 URL 中 & 等特殊字符。
    """
    npm_bin = Path.home() / "AppData" / "Roaming" / "npm"
    opencli_exe = str(npm_bin / "opencli.cmd") if sys.platform == "win32" else "opencli"
    is_list = "search" in args

    # 含特殊字符的参数加双引号
    quoted = []
    for a in args:
        if any(c in a for c in ' &|<>()^;"'):
            quoted.append(f'"{a}"')
        else:
            quoted.append(a)
    cmd_str = f'"{opencli_exe}" ' + " ".join(quoted)

    try:
        result = subprocess.run(
            cmd_str, capture_output=True, timeout=timeout,
            encoding="utf-8", errors="replace", shell=True,
        )
        stdout = (result.stdout or "").strip()
        stderr = (result.stderr or "").strip()

        # 检测风控拦截（判断 error 对象中的 code）
        if "SECURITY_BLOCK" in stdout or "SECURITY_BLOCK" in stderr:
            return {"_blocked": True}

        if result.returncode != 0:
            if stderr and "SECURITY_BLOCK" not in stderr:
                print(f"   ⚠️ OpenCLI: {stderr[:150]}")
            if not stdout:
                return [] if is_list else {}

        if not stdout:
            return [] if is_list else {}

        return json.loads(stdout)
    except subprocess.TimeoutExpired:
        print(f"   ⏰ 超时 ({timeout}s)")
        return [] if is_list else {}
    except json.JSONDecodeError:
        return [] if is_list else {}
    except Exception as e:
        print(f"   ❌ 调用失败: {e}")
        return [] if is_list else {}


def random_sleep(lo: float = 1.0, hi: float = 3.0):
    """随机等待"""
    time.sleep(lo + random.random() * (hi - lo))


# ============================================================
# 🔍 搜索笔记
# ============================================================

def search_notes(keyword: str, max_notes: int = 4) -> list[NoteCard]:
    """用 OpenCLI 搜索小红书笔记"""
    print(f"   🔍 搜索: 「{keyword}」...")
    data = run_opencli(["xiaohongshu", "search", keyword, "-f", "json"], timeout=30)

    if not data or not isinstance(data, list):
        print(f"   😢 无结果")
        return []

    notes = []
    for item in data[:max_notes]:
        url = item.get("url", "")
        note_id = ""
        for seg in ["/explore/", "/search_result/"]:
            if seg in url:
                note_id = url.split(seg)[-1].split("?")[0]
                break

        notes.append(NoteCard(
            note_id=note_id,
            title=item.get("title", ""),
            author=item.get("author", ""),
            like_count=parse_count(item.get("likes", 0)),
            url=url,
            published_at=item.get("published_at", ""),
            keyword_source=keyword,
        ))

    print(f"   ✅ 收录 {len(notes)} 条")
    return notes


# ============================================================
# 💬 抓取评论（含 fallback）
# ============================================================

def scrape_comments(note: NoteCard, limit: int = 40) -> list[CommentRow]:
    """抓取单条笔记评论，返回 CommentRow 列表"""
    data = run_opencli([
        "xiaohongshu", "comments", note.url,
        "--with-replies", "true",
        "--limit", str(limit),
        "-f", "json",
    ], timeout=60)

    # 被风控拦截
    if isinstance(data, dict) and data.get("_blocked"):
        return []  # 返回空列表让调用方做 fallback

    if not data or not isinstance(data, list):
        return []

    comments = []
    main_idx = 0
    for item in data:
        text = (item.get("text") or "").strip()
        if is_noise(text):
            continue
        is_reply = item.get("is_reply", False)
        if not is_reply:
            main_idx += 1
        comments.append(CommentRow(
            parent_id=f"M{main_idx}" if not is_reply else f"M{main_idx}",
            ctype="子级回复" if is_reply else "主评论",
            text=text,
            likes=parse_count(item.get("likes", 0)),
            post_time=item.get("time", "") or "",
            reply_to=item.get("reply_to", "") or "",
            note_url=note.url,
            note_title=note.title,
        ))

    main_c = sum(1 for c in comments if c.ctype == "主评论")
    sub_c = sum(1 for c in comments if c.ctype == "子级回复")
    if main_c + sub_c > 0:
        print(f"   ✅ {main_c} 主 + {sub_c} 子 = {len(comments)} 条")
    return comments


def scrape_with_fallback(notes: list[NoteCard], limit: int = 40) -> list[CommentRow]:
    """
    按热度降序尝试抓取评论，被风控就 fallback 到下一条。
    最多尝试 FALLBACK_DEPTH 条。
    """
    sorted_notes = sorted(notes, key=lambda n: n.like_count, reverse=True)

    for i, note in enumerate(sorted_notes[:FALLBACK_DEPTH]):
        if i == 0:
            print(f"   💬 抓评论: ❤️ {note.like_count} | {note.title[:40]}")
        else:
            print(f"   🔄 fallback [{i}]: ❤️ {note.like_count} | {note.title[:40]}")

        # 评论前额外休息
        random_sleep(COMMENT_NOTE_DELAY_MIN, COMMENT_NOTE_DELAY_MAX)

        result = scrape_comments(note, limit=limit)
        if result:
            return result
        else:
            print(f"   ⚠️ 被拦，尝试下一条...")

    print(f"   ❌ {FALLBACK_DEPTH} 条全被拦，跳过该关键词")
    return []


# ============================================================
# 📊 Excel 导出
# ============================================================

def export_excel(all_notes: list[NoteCard], all_comments: list[CommentRow]) -> str:
    """导出 Excel：笔记列表 + 评论明细（主评论浅紫色）"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = OUTPUT_DIR / f"饮料赛道_小红书_完整版_{timestamp}.xlsx"

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Sheet 1: 笔记列表
        if all_notes:
            pd.DataFrame([{
                "序号": i + 1,
                "笔记ID": n.note_id,
                "标题": n.title,
                "作者": n.author,
                "点赞数": n.like_count,
                "发布日期": n.published_at,
                "来源关键词": n.keyword_source,
                "链接": n.url,
            } for i, n in enumerate(all_notes)]).to_excel(
                writer, sheet_name="笔记列表", index=False)

        # Sheet 2: 评论明细
        if all_comments:
            df = pd.DataFrame([{
                "楼层ID": c.parent_id,
                "评论类型": c.ctype,
                "评论文本": c.text,
                "点赞数": c.likes,
                "发布时间": c.post_time,
                "被回复人": c.reply_to,
                "来源笔记": c.note_title,
                "笔记链接": c.note_url,
            } for c in all_comments])
            df.to_excel(writer, sheet_name="评论明细", index=False)

            # 主评论行浅紫色
            from openpyxl.styles import PatternFill
            ws = writer.sheets["评论明细"]
            purple = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")
            for row_idx in range(2, len(df) + 2):
                if ws.cell(row=row_idx, column=2).value == "主评论":
                    for col_idx in range(1, len(df.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = purple

    print(f"\n📊 Excel 已导出: {filepath}")
    print(f"   笔记: {len(all_notes)} 条 | 评论: {len(all_comments)} 条")
    return str(filepath)


# ============================================================
# 🎮 主流程
# ============================================================

def main():
    total_kw = sum(len(kws) for kws in KEYWORD_GROUPS.values())
    print("=" * 60)
    print("🧋 饮料赛道小红书抓取 v2")
    print("=" * 60)
    print(f"📂 {len(KEYWORD_GROUPS)} 组 × 每关键词 {MAX_NOTES_PER_KEYWORD} 笔记 = 预计 {total_kw * MAX_NOTES_PER_KEYWORD} 条")
    print(f"📂 评论：全局热度 TOP {TOP_COMMENT_NOTES} 笔记（笔记间 {COMMENT_NOTE_DELAY_MIN}~{COMMENT_NOTE_DELAY_MAX}s）")
    print("=" * 60)

    all_notes: list[NoteCard] = []
    all_comments: list[CommentRow] = []
    seen_ids: set = set()

    # ════════════════════════════════════
    # 🔍 阶段一：搜索全部关键词
    # ════════════════════════════════════
    for group_name, keywords in KEYWORD_GROUPS.items():
        print(f"\n{'─' * 50}")
        print(f"📂 【{group_name}】")
        print(f"{'─' * 50}")

        for kw in keywords:
            notes = search_notes(kw, max_notes=MAX_NOTES_PER_KEYWORD)
            new_notes = [n for n in notes if n.note_id and n.note_id not in seen_ids]
            for n in new_notes:
                seen_ids.add(n.note_id)
                all_notes.append(n)
            print(f"   📊 去重 {len(new_notes)} 条 | 累计 {len(all_notes)} 条")
            random_sleep(KEYWORD_DELAY_MIN, KEYWORD_DELAY_MAX)

    # ════════════════════════════════════
    # 💬 阶段二：全局热度 TOP N 集中抓评论
    # ════════════════════════════════════
    top_notes = sorted(all_notes, key=lambda n: n.like_count, reverse=True)[:TOP_COMMENT_NOTES]
    print(f"\n{'─' * 50}")
    print(f"💬 评论阶段：热度 TOP {len(top_notes)} 条笔记")
    print(f"{'─' * 50}")

    for i, note in enumerate(top_notes, 1):
        print(f"\n📌 [{i}/{len(top_notes)}] ❤️ {note.like_count} | {note.title[:45]}")
        print(f"   🔗 {note.url}")

        # 笔记之间长间隔
        if i > 1:
            delay = COMMENT_NOTE_DELAY_MIN + random.random() * (COMMENT_NOTE_DELAY_MAX - COMMENT_NOTE_DELAY_MIN)
            print(f"   ⏳ 等待 {delay:.1f}s...")
            time.sleep(delay)

        comments = scrape_with_fallback([note], limit=MAX_COMMENTS_PER_NOTE)
        if not comments:
            # fallback：从该笔记所属关键词组里找次热笔记
            related = [n for n in all_notes
                       if n.keyword_source == note.keyword_source and n.note_id != note.note_id]
            for fallback_note in sorted(related, key=lambda n: n.like_count, reverse=True)[:2]:
                print(f"   🔄 同关键词 fallback: ❤️ {fallback_note.like_count} | {fallback_note.title[:40]}")
                time.sleep(random.uniform(3, 7))
                comments = scrape_comments(fallback_note, limit=MAX_COMMENTS_PER_NOTE)
                if comments:
                    break

        all_comments.extend(comments)

    # ── 导出 ──
    if all_notes:
        export_excel(all_notes, all_comments)
    else:
        print("\n😢 什么也没抓到！检查 opencli doctor 和 Chrome 登录态~")
        return

    # ── 摘要 ──
    from collections import Counter
    print("\n" + "=" * 60)
    print("🎉 完成！")
    print("=" * 60)

    kw_stats = Counter(n.keyword_source for n in all_notes)
    print("\n📊 笔记来源:")
    for kw, cnt in kw_stats.most_common():
        print(f"   「{kw}」: {cnt} 条")

    top5 = sorted(all_notes, key=lambda n: n.like_count, reverse=True)[:5]
    print(f"\n🔥 TOP 5:")
    for i, n in enumerate(top5, 1):
        print(f"   {i}. ❤️ {n.like_count:>6} | {n.title[:50]}")

    main_c = sum(1 for c in all_comments if c.ctype == "主评论")
    sub_c = sum(1 for c in all_comments if c.ctype == "子级回复")
    print(f"\n💬 评论: {len(all_comments)} 条（{main_c} 主 + {sub_c} 子）")
    print(f"\n📁 output/ 目录下吼~")


if __name__ == "__main__":
    main()
