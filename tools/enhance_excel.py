"""
Excel 增强脚本 —— 给抓取结果做归类整理
==========================================
改进点：
  1. 笔记列表 → 新增「组别」列，按组排序
  2. 评论明细 → 新增「组别」列
  3. 新增「数据概览」Sheet → 各组统计
  4. 新增「分组评论」Sheet → 按组分区，组间标题行
  5. 主评论浅紫色填充保留
  6. 自动筛选 + 列宽自适应
"""

import pandas as pd
import io
import os
import sys
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 项目根目录
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# UTF-8 输出
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ════════════════════════════════════
# 配置
# ════════════════════════════════════
# 自动找最新的完整版 Excel
OUTPUT_DIR = PROJECT_ROOT / "output"
full_versions = sorted(OUTPUT_DIR.glob("饮料赛道_小红书_完整版_*.xlsx"))
if not full_versions:
    print("❌ 找不到完整版 Excel！请先运行 beverage_scanner_v2.py")
    sys.exit(1)
SRC = full_versions[-1]  # 最新的
DST = OUTPUT_DIR / f"饮料赛道_小红书_增强版_{datetime.now().strftime('%Y%m%d')}.xlsx"

# 关键词 → 组别 映射（从 config 读取）
try:
    from config import KW_TO_GROUP, GROUP_ORDER
except ImportError:
    print("❌ 找不到 config.py！")
    print("   请先复制 config.example.py → config.py 并填入你的关键词")
    sys.exit(1)

# 颜色定义
PURPLE_FILL = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
GROUP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # 组标题行浅蓝
GROUP_FONT = Font(name="微软雅黑", size=12, bold=True, color="1F3864")
NORMAL_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F3864")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

print("📖 读取原始数据...")
notes = pd.read_excel(SRC, sheet_name="笔记列表")
comments = pd.read_excel(SRC, sheet_name="评论明细")

# ── 添加组别列 ──
notes["组别"] = notes["来源关键词"].map(KW_TO_GROUP)
# 按组别排序
notes["_sort"] = notes["组别"].apply(lambda g: GROUP_ORDER.index(g) if g in GROUP_ORDER else 99)
notes = notes.sort_values(["_sort", "点赞数"], ascending=[True, False]).drop(columns=["_sort"]).reset_index(drop=True)
notes["序号"] = range(1, len(notes) + 1)

# 评论也加组别（通过笔记标题匹配，可能有同名笔记但概率低）
note_group_map = dict(zip(notes["标题"], notes["组别"]))
comments["组别"] = comments["来源笔记"].map(note_group_map)
# 没匹配到的用关键词反查
comments["组别"] = comments["组别"].fillna("其他")

print(f"✅ 笔记 {len(notes)} 条 | 评论 {len(comments)} 条")

# ════════════════════════════════════
# 写入 Excel
# ════════════════════════════════════
print("📝 生成增强版 Excel...")

with pd.ExcelWriter(DST, engine="openpyxl") as writer:
    # ────── Sheet 1: 笔记列表 ──────
    note_cols = ["序号", "组别", "笔记ID", "标题", "作者", "点赞数", "发布日期", "来源关键词", "链接"]
    notes[note_cols].to_excel(writer, sheet_name="笔记列表", index=False)
    ws_notes = writer.sheets["笔记列表"]
    # 表头样式
    for ci, col in enumerate(note_cols, 1):
        cell = ws_notes.cell(row=1, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 数据行样式 + 组别着色
    for ri in range(2, len(notes) + 2):
        group_val = ws_notes.cell(row=ri, column=2).value
        for ci in range(1, len(note_cols) + 1):
            cell = ws_notes.cell(row=ri, column=ci)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if ci >= 3:
                cell.alignment = Alignment(vertical="center")
    # 自动筛选
    ws_notes.auto_filter.ref = f"A1:{get_column_letter(len(note_cols))}{len(notes)+1}"
    # 列宽
    col_widths_n = [6, 16, 28, 40, 16, 10, 14, 18, 50]
    for ci, w in enumerate(col_widths_n, 1):
        ws_notes.column_dimensions[get_column_letter(ci)].width = w
    ws_notes.freeze_panes = "A2"

    # ────── Sheet 2: 评论明细 ──────
    cmt_cols = ["组别", "楼层ID", "评论类型", "评论文本", "点赞数", "发布时间", "被回复人", "来源笔记", "笔记链接"]
    # 按组别排序评论
    comments["_sort"] = comments["组别"].apply(lambda g: GROUP_ORDER.index(g) if g in GROUP_ORDER else 99)
    comments = comments.sort_values(["_sort", "评论类型", "点赞数"], ascending=[True, False, False]).drop(columns=["_sort"]).reset_index(drop=True)
    comments[cmt_cols].to_excel(writer, sheet_name="评论明细", index=False)
    ws_cmt = writer.sheets["评论明细"]
    # 表头样式
    for ci, col in enumerate(cmt_cols, 1):
        cell = ws_cmt.cell(row=1, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 数据行 + 主评论紫色
    for ri in range(2, len(comments) + 2):
        ctype = ws_cmt.cell(row=ri, column=3).value
        for ci in range(1, len(cmt_cols) + 1):
            cell = ws_cmt.cell(row=ri, column=ci)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if ci >= 4:
                cell.alignment = Alignment(vertical="center")
            # 主评论浅紫色填充
            if ctype == "主评论":
                cell.fill = PURPLE_FILL
    ws_cmt.auto_filter.ref = f"A1:{get_column_letter(len(cmt_cols))}{len(comments)+1}"
    col_widths_c = [16, 8, 10, 60, 8, 14, 14, 40, 50]
    for ci, w in enumerate(col_widths_c, 1):
        ws_cmt.column_dimensions[get_column_letter(ci)].width = w
    ws_cmt.freeze_panes = "A2"

    # ────── Sheet 3: 数据概览 ──────
    summary_rows = []
    for g in GROUP_ORDER:
        gn = notes[notes["组别"] == g]
        gc = comments[comments["组别"] == g]
        main_c = len(gc[gc["评论类型"] == "主评论"])
        sub_c = len(gc[gc["评论类型"] == "子级回复"])
        summary_rows.append({
            "组别": g,
            "笔记数": len(gn),
            "评论总数": len(gc),
            "主评论": main_c,
            "子回复": sub_c,
            "平均点赞": round(gn["点赞数"].mean(), 0) if len(gn) > 0 else 0,
            "最高点赞": gn["点赞数"].max() if len(gn) > 0 else 0,
            "最低点赞": gn["点赞数"].min() if len(gn) > 0 else 0,
        })
    # 合计行
    summary_rows.append({
        "组别": "📊 合计",
        "笔记数": len(notes),
        "评论总数": len(comments),
        "主评论": sum(1 for _, r in comments.iterrows() if r["评论类型"] == "主评论"),
        "子回复": sum(1 for _, r in comments.iterrows() if r["评论类型"] == "子级回复"),
        "平均点赞": round(notes["点赞数"].mean(), 0),
        "最高点赞": notes["点赞数"].max(),
        "最低点赞": notes["点赞数"].min(),
    })
    df_sum = pd.DataFrame(summary_rows)
    df_sum.to_excel(writer, sheet_name="数据概览", index=False)
    ws_sum = writer.sheets["数据概览"]
    sum_cols = list(df_sum.columns)
    for ci in range(1, len(sum_cols) + 1):
        cell = ws_sum.cell(row=1, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri in range(2, len(summary_rows) + 2):
        for ci in range(1, len(sum_cols) + 1):
            cell = ws_sum.cell(row=ri, column=ci)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
    # 合计行加粗
    last_row = len(summary_rows) + 1
    for ci in range(1, len(sum_cols) + 1):
        ws_sum.cell(row=last_row, column=ci).font = Font(name="微软雅黑", size=10, bold=True)
    sum_widths = [18, 10, 10, 10, 10, 12, 12, 12]
    for ci, w in enumerate(sum_widths, 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.freeze_panes = "A2"

    # ────── Sheet 4: 分组评论（按组分区，组间标题行） ──────
    # 这个 Sheet 手动写入，不用 pandas
    ws_gc = writer.book.create_sheet("分组评论")
    current_row = 1
    # 列定义
    gc_cols = ["楼层ID", "评论类型", "评论文本", "点赞数", "发布时间", "被回复人", "来源笔记"]
    gc_widths = [8, 10, 65, 8, 14, 14, 42]

    for gi, group_name in enumerate(GROUP_ORDER):
        gc = comments[comments["组别"] == group_name]
        if len(gc) == 0:
            continue

        # 组标题行
        ws_gc.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(gc_cols))
        title_cell = ws_gc.cell(row=current_row, column=1)
        title_cell.value = f"▍{group_name}（{len(gc)} 条评论）"
        title_cell.font = GROUP_FONT
        title_cell.fill = GROUP_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        for ci in range(1, len(gc_cols) + 1):
            ws_gc.cell(row=current_row, column=ci).fill = GROUP_FILL  # 合并区域的其余格也填色
        ws_gc.row_dimensions[current_row].height = 28
        current_row += 1

        # 表头（每组一个表头）
        for ci, col_name in enumerate(gc_cols, 1):
            cell = ws_gc.cell(row=current_row, column=ci)
            cell.value = col_name
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # 数据行
        for _, row in gc.iterrows():
            for ci, col_name in enumerate(gc_cols, 1):
                cell = ws_gc.cell(row=current_row, column=ci)
                cell.value = row[col_name]
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                if col_name == "评论文本":
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="center")
                # 主评论浅紫色
                if row["评论类型"] == "主评论":
                    cell.fill = PURPLE_FILL
            current_row += 1

        # 组间空行
        current_row += 1

    # 列宽
    for ci, w in enumerate(gc_widths, 1):
        ws_gc.column_dimensions[get_column_letter(ci)].width = w
    ws_gc.freeze_panes = "A2"

print(f"💾 保存到: {DST}")
print("🎉 增强版 Excel 生成完毕！")
